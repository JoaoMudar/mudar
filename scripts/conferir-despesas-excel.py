#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Conferencia linha a linha: financeiro.despesas x planilhas DESPESAS 20xx.xls

SOMENTE LEITURA. Nao escreve no banco, nao altera as planilhas.

Por que existe
--------------
O historico de despesas foi importado das 24 planilhas anuais para um banco
externo (`notas_despesas`) e de la veio para `financeiro.despesas` pelo
`scripts/import-financeiro.ps1`. O script que leu os Excel se perdeu
(`migracao/import_despesas.py` esta vazio), entao nao havia prova de que o banco
reflete a planilha.

A `financeiro.vw_bi_conferencia_mensal` compara o totalizador da planilha com a
soma do detalhe — mas os dois lados vieram do MESMO import. Ela pega linha
perdida dentro de um mes; nao pega valor trocado, data errada, descricao
truncada nem mes inteiro faltando.

Este script le os .xls originais e compara campo a campo. E exato (nao
amostral) porque a tabela guarda a proveniencia de cada linha: `fonte`
(arquivo), `aba` (nome da aba) e `linha_excel` (linha 1-based do Excel).

Uso
---
    npm run db:conferir-despesas
    npm run db:conferir-despesas -- --ano 2016
    npm run db:conferir-despesas -- --self-test

Sai com codigo != 0 se houver qualquer divergencia.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
import tempfile
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
DIR_PLANILHAS_PADRAO = RAIZ.parent / 'migracao' / 'dados' / 'despesas'

# `fonte` no banco -> arquivo real no disco. O 2004 foi importado sob o nome
# "novo"; conferido a mao, o conteudo bate com o DESPESAS 2004.xlsx da pasta.
ALIAS_FONTE = {
    'DESPESAS 2004 novo.xlsx': 'DESPESAS 2004.xlsx',
}

# Colunas A..I da planilha, na ordem, e a coluna correspondente no banco.
COLUNAS = [
    ('data',         'data'),
    ('quantidade',   'quantidade'),
    ('unidade',      'unidade'),
    ('descricao',    'descricao'),
    ('local',        'local'),
    ('valor_mc',     'valor_mc'),
    ('mao_obra',     'mao_obra'),
    ('equipamento',  'equipamento'),
    ('deslocamento', 'deslocamento'),
]
CAMPOS = [c for c, _ in COLUNAS]
CAMPOS_DINHEIRO = {'valor_mc', 'mao_obra', 'equipamento', 'deslocamento'}
CAMPOS_TEXTO = {'unidade', 'descricao', 'local'}
# Campos que o import normalizou para minusculas — comparar sem olhar a caixa.
CAMPOS_CAIXA_LIVRE = {'unidade', 'local'}

# Layout tipico: linhas 1-2 sao titulo, linha 3 e o cabecalho, dados a partir da 4.
# Mas algumas abas (Out08, Nov08, Dez08, Out09..Dez10) nao tem as linhas de titulo
# e o cabecalho esta na linha 1 — por isso a linha de inicio e detectada, nao fixa.
PRIMEIRA_LINHA_DADOS = 4
LINHAS_BUSCA_CABECALHO = 6
MESES_RE = re.compile(r'^(Jan|Fev|Mar|Abr|Mai|Jun|Jul|Ago|Set|Out|Nov|Dez)\s*\d{0,2}$', re.I)

TOL_DINHEIRO = Decimal('0.01')
TOL_QUANTIDADE = Decimal('0.001')


# ---------------------------------------------------------------------------
# Normalizacao — funcoes puras, cobertas pelo --self-test
# ---------------------------------------------------------------------------

def norm_texto(valor):
    """Texto da planilha -> forma canonica. Remove o `*` que prefixa toda
    descricao no Excel e colapsa espacos. Vazio vira None."""
    if valor is None:
        return None
    s = str(valor).strip()
    if s.startswith('*'):
        s = s[1:]
    s = re.sub(r'\s+', ' ', s).strip()
    return s or None


def norm_numero(valor, casas=2):
    """Numero da planilha -> Decimal arredondado. O Excel guarda 12191.659999999998
    onde o banco tem 12191.66, entao arredondar e obrigatorio."""
    if valor is None or valor == '':
        return None
    if isinstance(valor, str):
        s = valor.strip().replace('.', '').replace(',', '.') if ',' in valor else valor.strip()
        if not s:
            return None
        try:
            valor = float(s)
        except ValueError:
            return s  # texto numa coluna numerica: reporta como divergencia
    return Decimal(str(round(float(valor), casas)))


def norm_data(valor, datemode=0):
    """Serial do Excel / datetime -> date. Vazio vira None."""
    if valor is None or valor == '':
        return None
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    if isinstance(valor, (int, float)):
        import xlrd
        if float(valor) <= 0:
            return None
        try:
            return xlrd.xldate.xldate_as_datetime(float(valor), datemode).date()
        except Exception:
            return None
    s = str(valor).strip()
    return s or None


def acha_cabecalho(primeiras_linhas):
    """Devolve a linha 1-based onde comecam os dados.

    Procura o cabecalho `DATA | QUANT. | UNID. | DESCRICAO ...` nas primeiras
    linhas e devolve a linha seguinte. Sem cabecalho, cai no layout padrao.
    """
    for i, celulas in enumerate(primeiras_linhas, start=1):
        primeira = str(celulas[0] or '').strip().upper()
        segunda = str(celulas[1] or '').strip().upper() if len(celulas) > 1 else ''
        if primeira.startswith('DATA') and segunda.startswith('QUANT'):
            return i + 1
    return PRIMEIRA_LINHA_DADOS


def linha_vazia(linha):
    """Linha totalmente em branco nas 9 colunas — foi o que o import descartou."""
    return all(linha.get(c) is None for c in CAMPOS)


def valores_batem(campo, a, b):
    """Compara um campo dos dois lados. Devolve (igual, motivo)."""
    if a is None and b is None:
        return True, None
    if a is None or b is None:
        if campo in CAMPOS_DINHEIRO or campo == 'quantidade':
            outro = b if a is None else a
            if isinstance(outro, Decimal) and outro == 0:
                return False, 'vazio_vs_zero'
        return False, 'ausente_de_um_lado'
    if campo in CAMPOS_DINHEIRO:
        if not isinstance(a, Decimal) or not isinstance(b, Decimal):
            return str(a) == str(b), 'tipo'
        return abs(a - b) <= TOL_DINHEIRO, 'valor'
    if campo == 'quantidade':
        if not isinstance(a, Decimal) or not isinstance(b, Decimal):
            return str(a) == str(b), 'tipo'
        return abs(a - b) <= TOL_QUANTIDADE, 'valor'
    if campo == 'data':
        return a == b, 'data'
    if campo in CAMPOS_CAIXA_LIVRE:
        # O import padronizou centro de custo e unidade em minusculas
        # ("Campo" -> "campo"); diferenca de caixa nao e erro de importacao.
        return str(a).lower() == str(b).lower(), 'texto'
    return str(a) == str(b), 'texto'


def eh_totalizador_derivado(linha):
    """Regra re-derivada: rodape da aba nao tem descricao, so valor."""
    if linha.get('descricao') is not None:
        return False
    return any(linha.get(c) not in (None,) for c in CAMPOS_DINHEIRO)


# ---------------------------------------------------------------------------
# Leitura das planilhas
# ---------------------------------------------------------------------------

def ler_xls(caminho):
    """.xls BIFF (xlrd). Devolve {aba: {linha_excel: {campo: valor}}}."""
    import xlrd
    livro = xlrd.open_workbook(str(caminho))
    out = {}
    for nome in livro.sheet_names():
        if not MESES_RE.match(nome.strip()):
            continue
        aba = livro.sheet_by_name(nome)
        cabecalho = [aba.row_values(i) for i in range(min(LINHAS_BUSCA_CABECALHO, aba.nrows))]
        inicio = acha_cabecalho(cabecalho)
        linhas = {}
        for idx in range(inicio - 1, aba.nrows):
            celulas = aba.row_values(idx)
            celulas = (celulas + [''] * 9)[:9]
            linha = _monta_linha(celulas, livro.datemode)
            if linha_vazia(linha):
                continue
            linhas[idx + 1] = linha
        out[nome] = linhas
    return out


def ler_xlsx(caminho):
    """.xlsx OOXML (openpyxl). Mesmo formato de saida do ler_xls."""
    import openpyxl
    livro = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
    out = {}
    for nome in livro.sheetnames:
        if not MESES_RE.match(nome.strip()):
            continue
        aba = livro[nome]
        todas = [list(r) for r in aba.iter_rows(max_col=9, values_only=True)]
        inicio = acha_cabecalho(todas[:LINHAS_BUSCA_CABECALHO])
        linhas = {}
        for i, celulas in enumerate(todas[inicio - 1:], 0):
            celulas = (celulas + [None] * 9)[:9]
            linha = _monta_linha(celulas, 0)
            if linha_vazia(linha):
                continue
            linhas[inicio + i] = linha
        out[nome] = linhas
    livro.close()
    return out


def _monta_linha(celulas, datemode):
    return {
        'data':         norm_data(celulas[0], datemode),
        'quantidade':   norm_numero(celulas[1], 3),
        'unidade':      norm_texto(celulas[2]),
        'descricao':    norm_texto(celulas[3]),
        'local':        norm_texto(celulas[4]),
        'valor_mc':     norm_numero(celulas[5]),
        'mao_obra':     norm_numero(celulas[6]),
        'equipamento':  norm_numero(celulas[7]),
        'deslocamento': norm_numero(celulas[8]),
    }


def ler_planilha(caminho):
    return ler_xlsx(caminho) if caminho.suffix.lower() == '.xlsx' else ler_xls(caminho)


# ---------------------------------------------------------------------------
# Banco
# ---------------------------------------------------------------------------

def carrega_env():
    """Mesmo parsing manual do scripts/bi-sanity.ts (tsx/python nao leem .env sozinhos)."""
    env = RAIZ / '.env.local'
    if not env.exists():
        return
    for linha in env.read_text(encoding='utf-8').splitlines():
        linha = linha.strip()
        if not linha or linha.startswith('#') or '=' not in linha:
            continue
        chave, _, valor = linha.partition('=')
        os.environ.setdefault(chave, valor)


def carrega_banco(ano=None):
    """{(fonte, aba, linha_excel): {campo: valor, '_totalizador': bool}}"""
    import psycopg2
    url = os.environ.get('DATABASE_URL')
    if not url:
        sys.exit('DATABASE_URL nao definida (.env.local ou ambiente).')

    sql = """
        SELECT fonte, aba, linha_excel, data, quantidade, unidade, descricao,
               local, valor_mc, mao_obra, equipamento, deslocamento, eh_totalizador
          FROM financeiro.despesas
         WHERE excluido_em IS NULL
           AND origem_lancamento = 'excel'
           AND linha_excel IS NOT NULL
    """
    params = []
    if ano:
        sql += ' AND ano = %s'
        params.append(ano)

    conn = psycopg2.connect(url)
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            linhas = cur.fetchall()
    finally:
        conn.close()

    out = {}
    for r in linhas:
        fonte, aba, linha_excel = r[0], r[1], r[2]
        out[(fonte, aba, linha_excel)] = {
            'data':         r[3],
            'quantidade':   None if r[4] is None else Decimal(str(round(float(r[4]), 3))),
            'unidade':      norm_texto(r[5]),
            'descricao':    norm_texto(r[6]),
            'local':        norm_texto(r[7]),
            'valor_mc':     None if r[8] is None else Decimal(str(round(float(r[8]), 2))),
            'mao_obra':     None if r[9] is None else Decimal(str(round(float(r[9]), 2))),
            'equipamento':  None if r[10] is None else Decimal(str(round(float(r[10]), 2))),
            'deslocamento': None if r[11] is None else Decimal(str(round(float(r[11]), 2))),
            '_totalizador': r[12],
        }
    return out


# ---------------------------------------------------------------------------
# Comparacao
# ---------------------------------------------------------------------------

def confere(dir_planilhas, ano=None):
    banco = carrega_banco(ano)

    fontes = sorted({k[0] for k in banco})
    if ano:
        fontes = [f for f in fontes if str(ano) in f]

    achados = []     # divergencias campo a campo, viram CSV
    resumo = {}      # por fonte
    avisos = []

    for fonte in fontes:
        arquivo = dir_planilhas / ALIAS_FONTE.get(fonte, fonte)
        linhas_banco = {(a, l): v for (f, a, l), v in banco.items() if f == fonte}

        if not arquivo.exists():
            avisos.append(f'{fonte}: arquivo nao encontrado em {arquivo}')
            resumo[fonte] = dict(excel=0, banco=len(linhas_banco), iguais=0, divergentes=0,
                                 faltando=0, sobrando=0, dif=Decimal(0), sem_arquivo=True)
            continue

        planilha = ler_planilha(arquivo)
        linhas_excel = {(aba, ln): v for aba, ls in planilha.items() for ln, v in ls.items()}

        abas_banco = {a for a, _ in linhas_banco}
        abas_excel = set(planilha)
        for a in sorted(abas_banco - abas_excel):
            avisos.append(f'{fonte}: aba "{a}" existe no banco e nao no arquivo')
        for a in sorted(abas_excel - abas_banco):
            avisos.append(f'{fonte}: aba "{a}" existe no arquivo e nao no banco')

        contagem = defaultdict(int)
        dif_total = Decimal(0)

        for aba in sorted({a for a, _ in linhas_excel} | {a for a, _ in linhas_banco}):
            ex_aba = {ln: v for (a, ln), v in linhas_excel.items() if a == aba}
            bd_aba = {ln: v for (a, ln), v in linhas_banco.items() if a == aba}
            dif_total += confere_aba(fonte, aba, ex_aba, bd_aba, achados, contagem)

        resumo[fonte] = dict(
            excel=len(linhas_excel), banco=len(linhas_banco),
            iguais=contagem['iguais'], deslocadas=contagem['deslocadas'],
            divergentes=contagem['divergentes'], faltando=contagem['faltando'],
            sobrando=contagem['sobrando'], dif=dif_total, sem_arquivo=False,
        )

    return resumo, achados, avisos


def confere_aba(fonte, aba, ex_aba, bd_aba, achados, contagem):
    """Compara uma aba em duas passadas e devolve a diferenca em R$ da aba.

    1a passada, por `linha_excel`: e a conferencia que interessa, porque o
       import gravou o numero da linha. So pareia se as descricoes conferirem
       (ou se as duas forem vazias, caso dos rodapes) — sem isso, uma planilha
       que ganhou uma linha no meio faria o script comparar lancamentos que nao
       tem nada a ver um com o outro.
    2a passada, por CONTEUDO (data + descricao + local): o que sobrou da 1a e
       pareado de novo ignorando a posicao. Sem isso, uma unica linha inserida
       na planilha depois do import desalinha tudo abaixo dela e vira centenas
       de falsas divergencias — foi o que aconteceu em Jan26. O que casa aqui
       so mudou de lugar; o que NAO casa e perda ou sobra de verdade.
    """
    dif = Decimal(0)
    resto_ex, resto_bd = dict(ex_aba), dict(bd_aba)
    pares = []

    for ln in sorted(set(ex_aba) & set(bd_aba)):
        ex, bd = ex_aba[ln], bd_aba[ln]
        if not _mesma_identidade(ex, bd):
            continue
        del resto_ex[ln], resto_bd[ln]
        if _diferencas(ex, bd):
            pares.append((ln, ln))
        else:
            contagem['iguais'] += 1

    novos, resto_ex, resto_bd = casa_por_conteudo(resto_ex, resto_bd)
    pares.extend(novos)

    for ln_ex, ln_bd in sorted(pares):
        ex, bd = ex_aba[ln_ex], bd_aba[ln_bd]
        diffs = _diferencas(ex, bd)
        if not diffs:
            contagem['deslocadas'] += 1
            achados.append(_achado(fonte, aba, ln_ex, 'deslocada', ex,
                                   motivo=f'banco na linha {ln_bd}',
                                   valor_excel=_fmt(ex)))
            continue
        contagem['divergentes'] += 1
        posicao = '' if ln_ex == ln_bd else f'banco na linha {ln_bd}'
        for campo, motivo in diffs:
            if campo == 'eh_totalizador':
                ve, vb = str(eh_totalizador_derivado(ex)), str(bd['_totalizador'])
            else:
                ve, vb = _v(ex[campo]), _v(bd[campo])
                if campo in CAMPOS_DINHEIRO:
                    dif += _num(ex[campo]) - _num(bd[campo])
            achados.append(_achado(fonte, aba, ln_ex, 'divergente', ex, campo=campo,
                                   motivo=(motivo + ' ' + posicao).strip(),
                                   valor_excel=ve, valor_banco=vb))

    for ln in sorted(resto_ex):
        ex = resto_ex[ln]
        dif += sum((_num(ex[c]) for c in CAMPOS_DINHEIRO), Decimal(0))
        contagem['faltando'] += 1
        achados.append(_achado(fonte, aba, ln, 'faltando_no_banco', ex,
                               valor_excel=_fmt(ex)))

    for ln in sorted(resto_bd):
        bd = resto_bd[ln]
        dif -= sum((_num(bd[c]) for c in CAMPOS_DINHEIRO), Decimal(0))
        contagem['sobrando'] += 1
        achados.append(_achado(fonte, aba, ln, 'sobrando_no_banco', bd,
                               valor_banco=_fmt(bd)))

    return dif


def _achado(fonte, aba, linha, tipo, linha_origem, campo='', motivo='',
            valor_excel='', valor_banco=''):
    """Um achado do relatorio. `descricao` e `rodape` viajam junto porque sao o
    que permite separar, depois, lancamento de verdade de linha de subtotal."""
    return dict(fonte=fonte, aba=aba, linha_excel=linha, tipo=tipo, campo=campo,
                motivo=motivo, descricao=_v(linha_origem['descricao']),
                rodape='sim' if linha_origem['descricao'] is None else '',
                valor_excel=valor_excel, valor_banco=valor_banco)


def _mesma_identidade(ex, bd):
    """As duas linhas descrevem o mesmo lancamento? A descricao e o que
    identifica; rodape (sem descricao dos dois lados) tambem casa."""
    de, db = ex['descricao'], bd['descricao']
    if de is None and db is None:
        return True
    if de is None or db is None:
        return False
    return de.lower() == db.lower()


def _diferencas(ex, bd):
    """Lista de (campo, motivo) em que as duas linhas nao batem."""
    diffs = []
    for campo in CAMPOS:
        ok, motivo = valores_batem(campo, ex[campo], bd[campo])
        if not ok:
            diffs.append((campo, motivo))
    if eh_totalizador_derivado(ex) != bd['_totalizador']:
        diffs.append(('eh_totalizador', 'classificacao'))
    return diffs


def chaves_conteudo(linha):
    """Chaves de identidade da linha, da mais estrita para a mais frouxa.
    Linha sem descricao (rodape) nao tem identidade — nao entra no pareamento."""
    desc = linha['descricao']
    if desc is None:
        return []
    desc = desc.lower()
    local = (linha['local'] or '').lower()
    return [(linha['data'], desc, local), (desc, local), (desc,)]


def casa_por_conteudo(resto_ex, resto_bd):
    """Pareia as sobras por conteudo, aceitando criterios cada vez mais frouxos.
    Devolve (pares, sobra_excel, sobra_banco)."""
    pares = []
    ex, bd = dict(resto_ex), dict(resto_bd)
    for nivel in range(3):
        indice = defaultdict(list)
        for ln, linha in sorted(bd.items()):
            chaves = chaves_conteudo(linha)
            if chaves:
                indice[chaves[nivel]].append(ln)
        for ln_ex, linha in sorted(ex.items()):
            chaves = chaves_conteudo(linha)
            if not chaves:
                continue
            fila = indice.get(chaves[nivel], [])
            while fila:
                ln_bd = fila.pop(0)
                if ln_bd in bd:
                    pares.append((ln_ex, ln_bd))
                    del ex[ln_ex]
                    del bd[ln_bd]
                    break
    return pares, ex, bd


def _num(x):
    return x if isinstance(x, Decimal) else Decimal(0)


def _v(x):
    return '' if x is None else str(x)


def _fmt(linha):
    return ' | '.join(f'{c}={_v(linha[c])}' for c in CAMPOS if linha[c] is not None)


# ---------------------------------------------------------------------------
# Relatorio
# ---------------------------------------------------------------------------

def imprime(resumo, achados, avisos, csv_path):
    print()
    print('CONFERENCIA financeiro.despesas x planilhas DESPESAS 20xx')
    print('=' * 100)
    print(f'{"arquivo":<26}{"excel":>7}{"banco":>7}{"iguais":>8}{"desloc":>8}'
          f'{"diverg":>8}{"faltam":>8}{"sobram":>8}{"dif R$":>14}')
    print('-' * 100)

    tot = defaultdict(int)
    dif_geral = Decimal(0)
    campos = ('excel', 'banco', 'iguais', 'deslocadas', 'divergentes', 'faltando', 'sobrando')
    for fonte in sorted(resumo):
        r = resumo[fonte]
        if r['sem_arquivo']:
            print(f'{fonte:<26}{"— arquivo nao encontrado —":>60}')
            continue
        print(f'{fonte:<26}{r["excel"]:>7}{r["banco"]:>7}{r["iguais"]:>8}{r["deslocadas"]:>8}'
              f'{r["divergentes"]:>8}{r["faltando"]:>8}{r["sobrando"]:>8}{r["dif"]:>14,.2f}')
        for k in campos:
            tot[k] += r[k]
        dif_geral += r['dif']

    print('-' * 100)
    print(f'{"TOTAL":<26}{tot["excel"]:>7}{tot["banco"]:>7}{tot["iguais"]:>8}{tot["deslocadas"]:>8}'
          f'{tot["divergentes"]:>8}{tot["faltando"]:>8}{tot["sobrando"]:>8}{dif_geral:>14,.2f}')
    print()
    print('"desloc" = linha identica, so mudou de posicao porque a planilha foi editada')
    print('depois do import. Nao e perda de dado.')
    print()

    if achados:
        motivos = defaultdict(int)
        for a in achados:
            if a['tipo'] == 'deslocada':
                continue
            base = a['motivo'].split(' banco na linha')[0]
            motivos[(a['tipo'], a['campo'], base)] += 1
        print('Divergencias por tipo:')
        for (tipo, campo, motivo), n in sorted(motivos.items(), key=lambda x: -x[1]):
            rot = tipo if not campo else f'{tipo}: {campo} ({motivo})'
            print(f'  {n:>7}  {rot}')
        print()
        imprime_dinheiro_faltante(achados)

    if avisos:
        print('Avisos:')
        for a in avisos:
            print(f'  ! {a}')
        print()

    problemas = tot['divergentes'] + tot['faltando'] + tot['sobrando']
    if problemas == 0:
        casadas = tot['iguais'] + tot['deslocadas']
        print(f'OK — as {casadas} linhas do banco batem com o Excel campo a campo.')
    else:
        print(f'{problemas} linha(s) com problema. Detalhe em: {csv_path}')
    return problemas


def imprime_dinheiro_faltante(achados):
    """A pergunta que importa: quanto dinheiro esta na planilha e nao no banco?

    Duas formas de perder valor, e as duas contam aqui:
      - a linha inteira nao entrou (`faltando_no_banco`);
      - a linha entrou mas com o valor vazio. E o caso dos lancamentos fixos,
        que o Gilberto deixa pre-digitados em vermelho no comeco do mes e so
        preenche o valor quando a conta chega. Se o import passou no meio, a
        descricao veio e o valor nao.
    """
    por_aba = defaultdict(lambda: [0, Decimal(0), 0, Decimal(0)])
    for a in achados:
        # Rodape (subtotal por centro + total geral) e o mesmo dinheiro do
        # detalhe contado de novo — somar isso multiplicaria o mes por 3.
        if a['rodape']:
            continue
        chave = (a['fonte'], a['aba'])
        if a['tipo'] == 'faltando_no_banco':
            valor = Decimal(0)
            for parte in a['valor_excel'].split(' | '):
                campo, _, bruto = parte.partition('=')
                if campo in CAMPOS_DINHEIRO:
                    try:
                        valor += Decimal(bruto)
                    except Exception:
                        pass
            if valor:
                por_aba[chave][0] += 1
                por_aba[chave][1] += valor
        elif a['tipo'] == 'divergente' and a['campo'] in CAMPOS_DINHEIRO:
            try:
                ex = Decimal(a['valor_excel'])
            except Exception:
                continue
            vazio_no_banco = a['valor_banco'] in ('', '0', '0.00')
            if ex > 0 and vazio_no_banco:
                por_aba[chave][2] += 1
                por_aba[chave][3] += ex

    if not por_aba:
        return
    print('DINHEIRO QUE ESTA NA PLANILHA E NAO ESTA NO BANCO')
    print(f'  {"aba":<28}{"linhas fora":>12}{"R$":>14}{"valor vazio":>13}{"R$":>14}')
    tot = [0, Decimal(0), 0, Decimal(0)]
    for (fonte, aba), v in sorted(por_aba.items(), key=lambda x: -(x[1][1] + x[1][3])):
        print(f'  {fonte[9:13] + " " + aba:<28}{v[0]:>12}{v[1]:>14,.2f}{v[2]:>13}{v[3]:>14,.2f}')
        for i in range(4):
            tot[i] += v[i]
    print(f'  {"TOTAL":<28}{tot[0]:>12}{tot[1]:>14,.2f}{tot[2]:>13}{tot[3]:>14,.2f}')
    print(f'  Somando as duas colunas: R$ {tot[1] + tot[3]:,.2f}')
    print()


def grava_csv(achados, caminho):
    campos = ['fonte', 'aba', 'linha_excel', 'tipo', 'campo', 'motivo', 'descricao',
              'rodape', 'valor_excel', 'valor_banco']
    with open(caminho, 'w', newline='', encoding='utf-8-sig') as fh:
        w = csv.DictWriter(fh, fieldnames=campos, delimiter=';')
        w.writeheader()
        for a in achados:
            w.writerow(a)


# ---------------------------------------------------------------------------
# Self-test das normalizacoes (roda sem banco e sem arquivo)
# ---------------------------------------------------------------------------

def self_test():
    casos = []

    def ok(nome, obtido, esperado):
        casos.append((nome, obtido == esperado, obtido, esperado))

    ok('serial 42394 -> 2016-01-25', norm_data(42394.0), dt.date(2016, 1, 25))
    ok('serial 37649 -> 2003-01-28', norm_data(37649.0), dt.date(2003, 1, 28))
    ok('datetime -> date', norm_data(dt.datetime(2004, 1, 15)), dt.date(2004, 1, 15))
    ok('data vazia -> None', norm_data(''), None)
    ok('serial 0 -> None', norm_data(0), None)

    ok('remove asterisco', norm_texto('*Luz viveiro '), 'Luz viveiro')
    ok('colapsa espacos', norm_texto('*Agua   Casa'), 'Agua Casa')
    ok('texto vazio -> None', norm_texto('   '), None)
    ok('preserva acento', norm_texto('*Água casa'), 'Água casa')

    ok('arredonda 2 casas', norm_numero(12191.659999999998), Decimal('12191.66'))
    ok('numero vazio -> None', norm_numero(''), None)
    ok('zero e zero', norm_numero(0.0), Decimal('0.0'))
    ok('quantidade 3 casas', norm_numero(20.0004, 3), Decimal('20.0'))

    ok('tolerancia de 1 centavo',
       valores_batem('valor_mc', Decimal('321.41'), Decimal('321.42'))[0], True)
    ok('2 centavos ja diverge',
       valores_batem('valor_mc', Decimal('321.41'), Decimal('321.44'))[0], False)
    ok('None x None batem', valores_batem('valor_mc', None, None)[0], True)
    ok('None x zero nao batem',
       valores_batem('valor_mc', None, Decimal('0'))[1], 'vazio_vs_zero')
    ok('datas iguais batem',
       valores_batem('data', dt.date(2016, 1, 25), dt.date(2016, 1, 25))[0], True)
    ok('local ignora caixa', valores_batem('local', 'Campo', 'campo')[0], True)
    ok('descricao respeita caixa',
       valores_batem('descricao', 'Luz Viveiro', 'luz viveiro')[0], False)

    vazia = {c: None for c in CAMPOS}
    ok('linha toda vazia e descartada', linha_vazia(vazia), True)
    com_valor = dict(vazia, valor_mc=Decimal('10'))
    ok('linha com valor nao e vazia', linha_vazia(com_valor), False)

    cab = [['CONTROLE DE DESPESAS'], ['JANEIRO', 2015], ['DATA ', 'QUANT.', 'UNID.']]
    ok('cabecalho na linha 3 -> dados na 4', acha_cabecalho(cab), 4)
    ok('cabecalho na linha 1 -> dados na 2 (Out08)',
       acha_cabecalho([['DATA ', 'QUANT.', 'UNID.'], [39749.0, '', '']]), 2)
    ok('sem cabecalho cai no padrao', acha_cabecalho([['x'], ['y']]), 4)

    ok('rodape sem descricao e totalizador', eh_totalizador_derivado(com_valor), True)
    lancamento = dict(vazia, descricao='Luz viveiro', valor_mc=Decimal('321.41'))
    ok('lancamento normal nao e totalizador', eh_totalizador_derivado(lancamento), False)

    # Pareamento por conteudo — o que salva o relatorio quando a planilha ganha
    # uma linha nova depois do import e tudo abaixo desce uma posicao.
    def lin(desc, data=None, local='casa', mc=None):
        return dict({c: None for c in CAMPOS}, descricao=desc, data=data, local=local,
                    valor_mc=mc, _totalizador=False)

    ex = {10: lin('Estacionamento'), 11: lin('Super Kiko'), 12: lin('Diesel RAM')}
    bd = {10: lin('Super Kiko'), 11: lin('Diesel RAM')}
    pares, sobra_ex, sobra_bd = casa_por_conteudo(ex, bd)
    ok('linha inserida: pareia os deslocados', sorted(pares), [(11, 10), (12, 11)])
    ok('linha inserida: sobra so a nova', list(sobra_ex), [10])
    ok('linha inserida: nada sobra no banco', list(sobra_bd), [])

    ex = {5: lin('Condomínio Saint Patrick', dt.date(2026, 4, 10), 'casa', Decimal('1432'))}
    bd = {5: lin('Condomínio Saint Patrick', dt.date(2026, 4, 10), 'casa', None)}
    pares, _, _ = casa_por_conteudo(ex, bd)
    ok('valor preenchido depois: mesma linha, ainda pareia', pares, [(5, 5)])
    ok('valor preenchido depois: acusa a diferenca',
       [c for c, _ in _diferencas(ex[5], bd[5])], ['valor_mc'])

    ok('rodape sem descricao nao entra no pareamento por conteudo',
       casa_por_conteudo({9: lin(None, mc=Decimal('100'))},
                         {9: lin(None, mc=Decimal('100'))})[0], [])
    ok('mas rodape casa pela linha', _mesma_identidade(lin(None), lin(None)), True)
    ok('descricao diferente na mesma linha nao e o mesmo lancamento',
       _mesma_identidade(lin('Super Kiko'), lin('Diesel RAM')), False)
    ok('caixa nao separa lancamento',
       _mesma_identidade(lin('Super Kiko'), lin('super kiko')), True)

    falhas = [c for c in casos if not c[1]]
    for nome, passou, obtido, esperado in casos:
        print(f'  {"ok  " if passou else "FALHA"}  {nome}'
              + ('' if passou else f'   obtido={obtido!r} esperado={esperado!r}'))
    print(f'\n{len(casos) - len(falhas)}/{len(casos)} verificacoes passaram.')
    return 1 if falhas else 0


# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description='Confere financeiro.despesas contra as planilhas.')
    p.add_argument('--dir', type=Path, default=DIR_PLANILHAS_PADRAO,
                   help='pasta das planilhas DESPESAS 20xx')
    p.add_argument('--ano', type=int, help='confere um ano so')
    p.add_argument('--out', type=Path,
                   default=Path(tempfile.gettempdir()) / 'conferencia-despesas.csv',
                   help='CSV com o detalhe das divergencias')
    p.add_argument('--allow-remote', action='store_true',
                   help='libera DATABASE_URL fora de localhost')
    p.add_argument('--self-test', action='store_true',
                   help='testa as normalizacoes, sem banco e sem arquivo')
    args = p.parse_args()

    if args.self_test:
        return self_test()

    carrega_env()
    url = os.environ.get('DATABASE_URL', '')
    if 'neon.tech' in url and not args.allow_remote:
        sys.exit('DATABASE_URL aponta para o Neon. Use --allow-remote se for proposital.')

    if not args.dir.is_dir():
        sys.exit(f'Pasta das planilhas nao encontrada: {args.dir}')

    resumo, achados, avisos = confere(args.dir, args.ano)
    if achados:
        grava_csv(achados, args.out)
    problemas = imprime(resumo, achados, avisos, args.out)
    return 1 if problemas else 0


if __name__ == '__main__':
    sys.exit(main())
